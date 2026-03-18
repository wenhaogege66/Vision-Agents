"""评审规则加载服务：扫描 rules/ 目录，加载并解析评审规则文件。

支持 .md, .pdf, .docx, .xlsx 格式的规则文件读取。
按 rules/{赛事}/{赛道}/{组别}/ 目录结构组织。
"""

import logging
import re
from pathlib import Path

from app.models.schemas import (
    CompetitionInfo,
    EvaluationDimension,
    EvaluationRules,
    GroupInfo,
    TrackInfo,
)

logger = logging.getLogger(__name__)

# ── 规则文件根目录 ────────────────────────────────────────────
# 相对于 backend/ 目录
_RULES_BASE = Path(__file__).resolve().parent.parent.parent / "rules"

# ── 支持的规则文件名（按优先级排序） ──────────────────────────
_SUPPORTED_FILES = ["rules.md", "rules.pdf", "rules.docx", "rules.xlsx"]

# ── 名称映射字典 ─────────────────────────────────────────────

COMPETITION_NAMES: dict[str, str] = {
    "guochuangsai": "中国国际大学生创新大赛（国创赛）",
    "datiao": "挑战杯大学生课外学术科技作品竞赛（大挑）",
    "xiaotiao": "挑战杯大学生创业计划竞赛（小挑）",
}

TRACK_NAMES: dict[str, str] = {
    "gaojiao": "高教主赛道",
    "honglv": "红旅赛道",
    "zhijiao": "职教赛道",
    "chanye": "产业命题赛道",
    "mengya": "萌芽赛道",
}

GROUP_NAMES: dict[str, str] = {
    "benke_chuangyi": "本科创意组",
    "benke_chuangye": "本科创业组",
    "yanjiusheng_chuangyi": "研究生创意组",
    "yanjiusheng_chuangye": "研究生创业组",
    "gongyi": "公益组",
    "chuangyi": "创意组",
    "chuangye": "创业组",
    "chengguo_zhuanhua": "成果转化组",
    "qiye_mingti": "企业命题组",
    "default": "默认组",
}


# ── 文件读取辅助函数 ──────────────────────────────────────────


def _read_md(path: Path) -> str:
    """读取 Markdown 文件内容"""
    return path.read_text(encoding="utf-8")


def _read_pdf(path: Path) -> str:
    """读取 PDF 文件文本内容"""
    try:
        from PyPDF2 import PdfReader

        reader = PdfReader(str(path))
        pages = [page.extract_text() or "" for page in reader.pages]
        return "\n".join(pages)
    except Exception as e:
        logger.warning("读取 PDF 文件失败: %s, 错误: %s", path, e)
        return ""


def _read_docx(path: Path) -> str:
    """读取 Word (.docx) 文件文本内容"""
    try:
        import docx

        doc = docx.Document(str(path))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n".join(paragraphs)
    except Exception as e:
        logger.warning("读取 DOCX 文件失败: %s, 错误: %s", path, e)
        return ""


def _read_xlsx(path: Path) -> str:
    """读取 Excel (.xlsx) 文件文本内容，将所有单元格拼接为文本"""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(path), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    lines.append("\t".join(cells))
        wb.close()
        return "\n".join(lines)
    except Exception as e:
        logger.warning("读取 XLSX 文件失败: %s, 错误: %s", path, e)
        return ""


# 扩展名 → 读取函数映射
_READERS: dict[str, callable] = {
    ".md": _read_md,
    ".pdf": _read_pdf,
    ".docx": _read_docx,
    ".xlsx": _read_xlsx,
}


def _read_rules_file(path: Path) -> str:
    """根据文件扩展名选择合适的读取函数"""
    reader = _READERS.get(path.suffix.lower())
    if reader is None:
        logger.warning("不支持的规则文件格式: %s", path.suffix)
        return ""
    return reader(path)


# ── 规则解析辅助函数 ──────────────────────────────────────────


def _parse_dimensions(raw_content: str) -> list[EvaluationDimension]:
    """从规则文本中解析评审维度。

    支持以下格式：
    1. Markdown 标题格式：## 维度名称（XX分）
    2. 括号格式：维度名称(XX分) 或 维度名称（XX分）
    3. Markdown 表格格式：| 维度名称 | XX |

    子项通过列表项（- 或 * 开头）或缩进行提取。
    """
    dimensions: list[EvaluationDimension] = []

    # 尝试按 Markdown 标题 + 列表项解析
    # 匹配 ## 维度名称（XX分） 或 ## 维度名称(XX分)
    heading_pattern = re.compile(
        r"^#{1,3}\s+(.+?)[（(]\s*(\d+)\s*分\s*[）)]",
        re.MULTILINE,
    )
    matches = list(heading_pattern.finditer(raw_content))

    if matches:
        for i, m in enumerate(matches):
            name = m.group(1).strip()
            max_score = float(m.group(2))

            # 提取该维度到下一个维度之间的子项
            start = m.end()
            end = matches[i + 1].start() if i + 1 < len(matches) else len(raw_content)
            section = raw_content[start:end]

            sub_items = _extract_sub_items(section)
            dimensions.append(
                EvaluationDimension(
                    name=name,
                    max_score=max_score,
                    sub_items=sub_items,
                )
            )
        return dimensions

    # 尝试按行内括号格式解析（非标题行）
    # 匹配：维度名称（XX分） 或 维度名称(XX分)
    inline_pattern = re.compile(
        r"^[*\-]?\s*(.+?)[（(]\s*(\d+)\s*分\s*[）)]",
        re.MULTILINE,
    )
    inline_matches = list(inline_pattern.finditer(raw_content))

    if inline_matches:
        for i, m in enumerate(inline_matches):
            name = m.group(1).strip().strip("*#- ")
            max_score = float(m.group(2))

            start = m.end()
            end = (
                inline_matches[i + 1].start()
                if i + 1 < len(inline_matches)
                else len(raw_content)
            )
            section = raw_content[start:end]

            sub_items = _extract_sub_items(section)
            dimensions.append(
                EvaluationDimension(
                    name=name,
                    max_score=max_score,
                    sub_items=sub_items,
                )
            )
        return dimensions

    # 尝试按 Markdown 表格解析
    # 匹配：| 维度名称 | XX | 子项1、子项2 |
    table_pattern = re.compile(
        r"^\|\s*(.+?)\s*\|\s*(\d+)\s*\|\s*(.*?)\s*\|",
        re.MULTILINE,
    )
    for m in table_pattern.finditer(raw_content):
        name = m.group(1).strip().strip("|").strip()
        # 跳过表头分隔行
        if name.startswith("-") or name.startswith(":"):
            continue
        # 跳过表头行（常见关键词）
        if name in ("维度", "评审维度", "维度名称", "名称"):
            continue
        try:
            max_score = float(m.group(2))
        except ValueError:
            continue
        sub_text = m.group(3).strip()
        sub_items = [s.strip() for s in re.split(r"[、,，;；]", sub_text) if s.strip()]
        dimensions.append(
            EvaluationDimension(
                name=name,
                max_score=max_score,
                sub_items=sub_items,
            )
        )

    return dimensions


def _extract_sub_items(section: str) -> list[str]:
    """从文本段落中提取子项列表（- 或 * 开头的行）"""
    sub_items: list[str] = []
    for line in section.splitlines():
        stripped = line.strip()
        if stripped.startswith(("- ", "* ", "· ")):
            item = stripped.lstrip("-*· ").strip()
            if item:
                sub_items.append(item)
    return sub_items


# ── RuleService 主类 ──────────────────────────────────────────


class RuleService:
    """评审规则加载服务。

    扫描 rules/ 目录结构，提供赛事/赛道/组别的列表查询，
    以及评审规则文件的加载和解析。
    """

    def __init__(self, rules_base: Path | None = None):
        """初始化规则服务。

        Args:
            rules_base: 规则文件根目录，默认为 backend/rules/
        """
        self._rules_base = rules_base or _RULES_BASE

    @property
    def rules_base(self) -> Path:
        """规则文件根目录"""
        return self._rules_base

    def _get_subdirs(self, parent: Path) -> list[str]:
        """获取目录下所有子目录名称（排除隐藏目录）"""
        if not parent.is_dir():
            return []
        return sorted(
            d.name
            for d in parent.iterdir()
            if d.is_dir() and not d.name.startswith(".")
        )

    def _find_rules_file(self, group_dir: Path) -> Path | None:
        """在组别目录中查找规则文件（按优先级）"""
        for filename in _SUPPORTED_FILES:
            path = group_dir / filename
            if path.is_file():
                return path
        return None

    def list_competitions(self) -> list[CompetitionInfo]:
        """列出所有可用赛事。

        扫描 rules/ 目录下的子目录，返回赛事列表。
        """
        competition_ids = self._get_subdirs(self._rules_base)
        return [
            CompetitionInfo(
                id=cid,
                name=COMPETITION_NAMES.get(cid, cid),
            )
            for cid in competition_ids
        ]

    def list_tracks(self, competition: str) -> list[TrackInfo]:
        """列出赛事下所有赛道。

        Args:
            competition: 赛事ID（如 guochuangsai）
        """
        competition_dir = self._rules_base / competition
        track_ids = self._get_subdirs(competition_dir)
        return [
            TrackInfo(
                id=tid,
                name=TRACK_NAMES.get(tid, tid),
            )
            for tid in track_ids
        ]

    def list_groups(self, competition: str, track: str) -> list[GroupInfo]:
        """列出赛道下所有组别。

        Args:
            competition: 赛事ID
            track: 赛道ID
        """
        track_dir = self._rules_base / competition / track
        group_ids = self._get_subdirs(track_dir)
        return [
            GroupInfo(
                id=gid,
                name=GROUP_NAMES.get(gid, gid),
                has_rules=self.has_rules(competition, track, gid),
            )
            for gid in group_ids
        ]

    def has_rules(self, competition: str, track: str, group: str) -> bool:
        """检查指定赛事/赛道/组别组合是否存在规则文件。

        Args:
            competition: 赛事ID
            track: 赛道ID
            group: 组别ID
        """
        group_dir = self._rules_base / competition / track / group
        return self._find_rules_file(group_dir) is not None

    def load_rules(
        self, competition: str, track: str, group: str
    ) -> EvaluationRules:
        """加载并解析指定赛事/赛道/组别的评审规则。

        Args:
            competition: 赛事ID
            track: 赛道ID
            group: 组别ID

        Returns:
            解析后的评审规则对象

        Raises:
            FileNotFoundError: 规则文件不存在
        """
        group_dir = self._rules_base / competition / track / group
        rules_file = self._find_rules_file(group_dir)

        if rules_file is None:
            raise FileNotFoundError(
                f"未找到评审规则文件: {competition}/{track}/{group}/"
            )

        # 读取文件内容
        raw_content = _read_rules_file(rules_file)
        if not raw_content.strip():
            raise FileNotFoundError(
                f"评审规则文件内容为空: {rules_file}"
            )

        # 解析评审维度
        dimensions = _parse_dimensions(raw_content)

        return EvaluationRules(
            competition=competition,
            track=track,
            group=group,
            dimensions=dimensions,
            raw_content=raw_content,
        )


# ── 模块级单例 ────────────────────────────────────────────────

rule_service = RuleService()
