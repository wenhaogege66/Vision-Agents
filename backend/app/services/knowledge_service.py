"""知识库加载服务：从 knowledge/ 目录读取评委经验知识库文件。

按材料类型（bp, text_ppt, presentation_ppt, presentation）分类组织，
支持 .md, .pdf, .docx, .xlsx 格式。
"""

import logging
from pathlib import Path

logger = logging.getLogger(__name__)

# 知识库根目录（相对于 backend/）
_KNOWLEDGE_BASE = Path(__file__).resolve().parent.parent.parent / "knowledge"

# 支持的文件扩展名
_SUPPORTED_EXTENSIONS = {".md", ".pdf", ".docx", ".xlsx"}

# 有效的材料类型子目录
VALID_MATERIAL_TYPES = {"bp", "text_ppt", "presentation_ppt", "presentation"}


def _read_file(path: Path) -> str:
    """根据扩展名读取文件内容，复用 rule_service 中的读取函数。"""
    ext = path.suffix.lower()
    if ext == ".md":
        return path.read_text(encoding="utf-8")
    elif ext == ".pdf":
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(str(path))
            return "\n".join(page.extract_text() or "" for page in reader.pages)
        except Exception as e:
            logger.warning("读取 PDF 失败: %s, 错误: %s", path, e)
            return ""
    elif ext == ".docx":
        try:
            import docx
            doc = docx.Document(str(path))
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except Exception as e:
            logger.warning("读取 DOCX 失败: %s, 错误: %s", path, e)
            return ""
    elif ext == ".xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(path), read_only=True, data_only=True)
            lines: list[str] = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        lines.append("\t".join(cells))
            wb.close()
            return "\n".join(lines)
        except Exception as e:
            logger.warning("读取 XLSX 失败: %s, 错误: %s", path, e)
            return ""
    else:
        logger.warning("不支持的知识库文件格式: %s", path.suffix)
        return ""


class KnowledgeService:
    """知识库加载服务。

    从 knowledge/{material_type}/ 目录读取所有支持格式的文件，
    拼接为完整的知识库文本内容。
    """

    def __init__(self, knowledge_base: Path | None = None):
        self._knowledge_base = knowledge_base or _KNOWLEDGE_BASE

    @property
    def knowledge_base(self) -> Path:
        return self._knowledge_base

    def load_knowledge(self, material_type: str) -> str:
        """加载指定材料类型的知识库内容。

        Args:
            material_type: 材料类型（bp / text_ppt / presentation_ppt / presentation）

        Returns:
            拼接后的知识库文本内容。目录不存在或无文件时返回空字符串。

        Raises:
            ValueError: material_type 不在有效范围内
        """
        if material_type not in VALID_MATERIAL_TYPES:
            raise ValueError(
                f"无效的材料类型: {material_type}，"
                f"有效值: {', '.join(sorted(VALID_MATERIAL_TYPES))}"
            )

        type_dir = self._knowledge_base / material_type
        if not type_dir.is_dir():
            logger.info("知识库目录不存在: %s", type_dir)
            return ""

        # 收集所有支持格式的文件，按文件名排序保证稳定顺序
        files = sorted(
            f for f in type_dir.iterdir()
            if f.is_file() and f.suffix.lower() in _SUPPORTED_EXTENSIONS
        )

        if not files:
            logger.info("知识库目录为空: %s", type_dir)
            return ""

        parts: list[str] = []
        for f in files:
            content = _read_file(f)
            if content.strip():
                parts.append(content)

        return "\n\n".join(parts)


# 模块级单例
knowledge_service = KnowledgeService()
